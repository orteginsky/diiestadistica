#qfdz csnf kkww irbg
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart


import smtplib
import os

def enviar_correo(
        ruta_carpeta,
        #correo_destinatario = "jalcaide@ipn.mx",
        correo_destinatario = "luisortegar99@gmail.com",
        correo_remitente = "diiestadisticabasica@gmail.com",
        contraseña_app = "qfdzcsnfkkwwirbg"
        ):
    ruta_archivo = f"{ruta_carpeta}/informes.zip"
    ruta_tabla = f"{ruta_carpeta}/archivo_maestro.xlsx"
    ruta_actual = os.path.dirname(os.path.abspath(__file__))
    ruta_firma = os.path.join(ruta_actual, "firma.html")

    with open(ruta_firma, "r", encoding="utf-8") as f:
        firma_html = f.read()

    mensaje_principal = """
    <p>Estimado/a,</p>
    <p>Le compartimos los archivos correspondientes. Favor de NO responder a este correo .</p>
    """

    # Destinatario y mensaje
    asunto = "Matricula Preeliminar"
    cuerpo = "Hola"

    # Crear mensaje
    mensaje = MIMEMultipart()
    mensaje["From"] = correo_remitente
    mensaje["To"] = correo_destinatario
    mensaje["Subject"] = asunto
    mensaje.attach(MIMEText(cuerpo, "plain"))

    # Combinar mensaje + firma
    html_final = mensaje_principal + firma_html

    # Adjuntar versión HTML
    parte_html = MIMEText(html_final, "html")
    mensaje.attach(parte_html)

    _, nombre_archivo = os.path.split(ruta_archivo)
    with open(ruta_archivo, "rb") as f:
        adjunto = MIMEApplication(f.read(), Name=nombre_archivo)
        adjunto['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        mensaje.attach(adjunto)

    # Enviar el correo
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as servidor:
        servidor.login(correo_remitente, contraseña_app)
        servidor.sendmail(mensaje["From"], mensaje["To"], mensaje.as_string())

    print("Correo enviado correctamente.")





