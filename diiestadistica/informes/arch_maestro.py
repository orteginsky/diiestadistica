from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
import numpy as np
import os

from diiestadistica.core.logging_config import setup_logger

logger = setup_logger(__name__)


def informes_mapre(ruta_carpeta):
    ruta_homo = os.path.normpath(os.path.join(ruta_carpeta,"archivos_homologados"))
    ruta_informe =  os.path.normpath(os.path.join(os.path.join(ruta_carpeta,"reportes"),"archivo_maestro.xlsx") )
    archivos_xlsx = [f for f in os.listdir(ruta_homo) if f.endswith(".xlsx")]
    
    if not archivos_xlsx:
        logger.warning("No se encontraron archivos .xlsx en la carpeta.")
        return None
    
    dataframes = [pd.read_excel(os.path.join(ruta_homo, archivo)) for archivo in archivos_xlsx]
    df_final = pd.concat(dataframes, ignore_index=True)
    df_final["Nivel"] = np.where(df_final["ID_Nivel"] == 1, "Medio Superior",
               np.where(df_final["ID_Nivel"] == 2, "Superior", "Posgrado"))

    df_final["Modalidad"] = np.where(
        df_final["Modalidad"] == 1,
        "Escolarizada",
        np.where(
            df_final["Modalidad"] == 2,
            "No escolarizada",
            "Mixta")
        )

    df_final["Rama"] = np.where(
        df_final["Rama"] == 1,
        "ICFM",
        np.where(
            df_final["Rama"] == 2,
            "CMB",
            np.where(
                df_final["Rama"] == 3,
                "CSA",
                np.where(
                    df_final["Rama"] == 4,
                    "CI",
                    "GENERAL")
                )
            )
        )
    #"""
    df_final = df_final[[
        "Periodo",
        "Nivel",
        "Indice",
        "Modalidad",
        "Rama",
        "Siglas",
        "id_Entidad_Federativa",
        "posgrado",
        "Programa",
        "Nombre_Programa",
        "Concepto",
        "Turno",
        "Sexo",
        "Fin_periodo",
        "Datos"]]
    #"""
    df_final.to_excel(ruta_informe,index=False)
    wb = load_workbook(ruta_informe)
    ws = wb.active

    encabezado_style = {
        "font": Font(size=12, bold=True, color="FFFFFF"),
        "fill": PatternFill("solid", fgColor="5A1236"),
        "alignment": Alignment(horizontal="center", vertical="center")
    }

    for col_idx, celda in enumerate(ws[1], 1):
        celda.font = encabezado_style["font"]
        celda.fill = encabezado_style["fill"]
        celda.alignment = encabezado_style["alignment"]

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(ruta_informe)



