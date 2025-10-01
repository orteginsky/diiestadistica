from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook

import pandas as pd
import re
import os

from diiestadistica.core.logging_config import setup_logger

logger = setup_logger(__name__)


def ordenar_y_agrupar_columna_en_libro(workbook: Workbook, nombre_columna: str):
    """
    Ordena y agrupa por una columna específica en todas las hojas de un Workbook de openpyxl.

    :param workbook: Objeto Workbook cargado con openpyxl
    :param nombre_columna: Nombre de la columna a ordenar y agrupar
    """
    for ws in workbook.worksheets:
        # Obtener encabezados
        encabezados = [cell.value for cell in ws[1]]

        if nombre_columna not in encabezados:
            continue

        col_idx = encabezados.index(nombre_columna) + 1
        col_letter = get_column_letter(col_idx)

        # Leer datos (sin encabezado)
        datos = list(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column, values_only=True))
        datos.sort(key=lambda fila: str(fila[col_idx-1]) if fila[col_idx-1] is not None else "")

        # Reescribir los datos ordenados
        for i, fila in enumerate(datos, start=2):
            for j, valor in enumerate(fila, start=1):
                ws.cell(row=i, column=j, value=valor)

        # Merge de celdas iguales en la columna objetivo
        fila_inicio = 2
        valor_anterior = ws[f"{col_letter}{fila_inicio}"].value

        for fila_actual in range(3, ws.max_row + 1):
            valor_actual = ws[f"{col_letter}{fila_actual}"].value
            if valor_actual != valor_anterior:
                if fila_actual - fila_inicio > 1:
                    ws.merge_cells(f"{col_letter}{fila_inicio}:{col_letter}{fila_actual - 1}")
                fila_inicio = fila_actual
                valor_anterior = valor_actual

        # Último grupo
        if ws.max_row + 1 - fila_inicio > 1:
            ws.merge_cells(f"{col_letter}{fila_inicio}:{col_letter}{ws.max_row}")


def anti_join(df1, df2, on):
    filas_repetidas = pd.merge(df1, df2[on], on=on, how='inner')
    condicion = ~df1.set_index(on).index.isin(filas_repetidas.set_index(on).index)
    return df1.loc[condicion].reset_index(drop=True)

def extraer_anio_de_periodo(periodo: str) -> int:
    """
    Extrae el año correspondiente de un string con formato 'periodo_XXXX-XXXX+1_Y'.
    Devuelve el primer o segundo año según el valor del semestre Y.

    :param periodo: String del tipo 'periodo_2024-2025_2'
    :return: Año como entero (ej. 2025)
    """
    try:
        base, semestre = periodo.split('_')[1], periodo.split('_')[2]
        anio_inicio, anio_fin = map(int, base.split('-'))
        return anio_fin if semestre == '2' else anio_inicio
    except Exception as e:
        logger.error(f"Error al extraer el año del periodo: {periodo}. Error: {e}")
        raise ValueError(f"Formato inválido del periodo: {periodo}. Error: {e}")

def filtrar_por_anio_disponible(df, columna_anio: str, anio: int):
    """
    Filtra el DataFrame por el año en la columna especificada.
    Si el año no está presente, intenta con el año anterior hasta encontrar coincidencias.

    :param df: DataFrame de pandas
    :param columna_anio: Nombre de la columna que contiene los años
    :param anio: Año base a buscar
    :return: DataFrame filtrado con el primer año encontrado hacia atrás
    """
    while anio >= df[columna_anio].min():
        df_filtrado = df[df[columna_anio] == anio]
        if not df_filtrado.empty:
            return df_filtrado
        anio -= 1
    return pd.DataFrame()  # Si no se encuentra ningún año válido


def pu(ruta_global):
    """
    Genera un libro de Excel con una hoja por cada valor único en la 'Indice'.
    En cada hoja realiza las una intersección con el catalogo de programas y unidades.
    Aplica formato a los encabezados.
    """
    ruta_reportes = os.path.normpath(os.path.join(ruta_global,"reportes"))
    ruta_archivo_maestro = os.path.normpath(os.path.join(ruta_reportes,"archivo_maestro.xlsx"))
    ruta_salida = os.path.normpath(os.path.join(ruta_reportes,"programas_no_reportados.xlsx"))
    ruta_pu = os.path.normpath(os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))),"historico.xlsx"))
    pu = pd.read_excel(ruta_pu)
    base_ruta = os.path.basename(ruta_global)
    anio_objetivo = extraer_anio_de_periodo(base_ruta)
    pu = filtrar_por_anio_disponible(pu,"ano",anio_objetivo)
    try:
        dataframe = pd.read_excel(ruta_archivo_maestro)
        indice = dataframe['Indice'].unique()
        hojas_generadas = []

        with pd.ExcelWriter(ruta_salida, engine='openpyxl', mode='w') as writer:
            for seleccion in indice:
                indice_dataframe = dataframe[dataframe['Indice'] == seleccion]
                agrupacion = indice_dataframe.groupby(['Siglas','Nombre_Programa','Modalidad'], as_index=False)['Datos'].sum()
                agrupacion = agrupacion[agrupacion['Datos']!=0]
                # hacer un anti join de PU - agrupacion
                npu = anti_join(pu,agrupacion,['Siglas','Nombre_Programa','Modalidad'])

                # Evita nombres de hoja inválidos
                nombre_hoja = re.sub(r'[\[\]\*\?\/\\]', '', seleccion)[:31]
                npu.to_excel(writer, sheet_name=nombre_hoja, index=False)
                hojas_generadas.append(nombre_hoja)
                
        # Aplicar estilos a los encabezados
        wb = load_workbook(ruta_salida)
        fill = PatternFill(start_color='5A1236', end_color='5A1236', fill_type='solid')
        font = Font(color='FFFFFF', bold=True, size=12)
        align = Alignment(horizontal='center', vertical='center')

        for hoja in hojas_generadas:
            ws = wb[hoja]
            for cell in ws[1]:
                cell.fill = fill
                cell.font = font
                cell.alignment = align

        ordenar_y_agrupar_columna_en_libro(wb, 'Siglas')
        wb.save(ruta_salida)
        
    except Exception as e:
        logger.info(f"Se produjo un error: {e}")


if __name__ == "__main__":
    logger.info(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

"""import pandas as pd

ruta_pu = "/home/kaliuser/Documentos/diiestadistica/historico.xlsx"

pu = pd.read_excel(ruta_pu)
logger.info(pu.columns)"""

"""def anti_join(df1, df2, left_on, right_on):
    # Hacemos un merge para encontrar las coincidencias
    filas_repetidas = pd.merge(
        df1,
        df2[right_on],
        left_on=left_on,
        right_on=right_on,
        how='inner'
    )

    # Convertimos a tuplas si son múltiples columnas
    index_df1 = df1[left_on].apply(tuple, axis=1) if isinstance(left_on, list) else df1[left_on]
    index_repetidas = filas_repetidas[left_on].apply(tuple, axis=1) if isinstance(left_on, list) else filas_repetidas[left_on]

    # Obtenemos lo que no está en el merge
    condicion = ~index_df1.isin(index_repetidas)

    return df1[condicion].reset_index(drop=True)"""