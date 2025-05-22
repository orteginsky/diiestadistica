from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

import pandas as pd
import re


def ordenar_y_agrupar_columna_en_libro(workbook: Workbook, nombre_columna: str, orden_deseado: list):
    """
    Ordena y agrupa por una columna específica en todas las hojas de un Workbook de openpyxl.

    :param workbook: Objeto Workbook cargado con openpyxl
    :param nombre_columna: Nombre de la columna a ordenar y agrupar
    :param orden_deseado: Lista con el orden deseado
    """
    for ws in workbook.worksheets:
        # Obtener encabezados
        encabezados = [cell.value for cell in ws[1]]

        if nombre_columna not in encabezados:
            continue

        col_idx = encabezados.index(nombre_columna) + 1
        col_letter = get_column_letter(col_idx)

        # Leer datos (sin encabezado)
        datos = list(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column, values_only=True))

        # Ordenar por el orden deseado
        datos.sort(key=lambda fila: orden_deseado.index(fila[col_idx - 1]) if fila[col_idx - 1] in orden_deseado else float('inf'))

        # Reescribir los datos ordenados
        for i, fila in enumerate(datos, start=2):
            for j, valor in enumerate(fila, start=1):
                ws.cell(row=i, column=j, value=valor)

        # Merge de celdas iguales en la columna objetivo
        fila_inicio = 2
        valor_anterior = ws[f"{col_letter}{fila_inicio}"].value

        for fila_actual in range(3, ws.max_row + 1):
            valor_actual = ws[f"{col_letter}{fila_actual}"].value
            if valor_actual != valor_anterior:
                if fila_actual - fila_inicio > 1:
                    ws.merge_cells(f"{col_letter}{fila_inicio}:{col_letter}{fila_actual - 1}")
                fila_inicio = fila_actual
                valor_anterior = valor_actual

        # Último grupo
        if ws.max_row + 1 - fila_inicio > 1:
            ws.merge_cells(f"{col_letter}{fila_inicio}:{col_letter}{ws.max_row}")


def mapre(ruta_global):
    """
    Genera un libro de Excel con una hoja por cada valor único en la 'Indice'.
    En cada hoja realiza las operaciones de agrupación, pivotado y creación de columnas adicionales.
    Aplica formato a los encabezados.
    """
    ruta_reportes = f"{ruta_global}/reportes"
    ruta_archivo_maestro = f"{ruta_reportes}/archivo_maestro.xlsx"
    ruta_salida = f"{ruta_reportes}/mapre.xlsx"
    #ruta_pu = 
    #pu = pd.read_excel(ruta_pu)
    #

    try:
        dataframe = pd.read_excel(ruta_archivo_maestro)
        indice = dataframe['Indice'].unique()
        hojas_generadas = []

        with pd.ExcelWriter(ruta_salida, engine='openpyxl', mode='w') as writer:
            if 'Grupos' in indice:
                indice_dataframe = dataframe[dataframe['Indice'] == 'Grupos']
                agrupacion = indice_dataframe.groupby(['Nivel', 'Sexo'], as_index=False)['Datos'].sum()
                agrupacion = agrupacion.pivot(index='Nivel', columns='Sexo', values='Datos').reset_index()
                agrupacion['Total'] = agrupacion.get('Hombres', 0).fillna(0) + agrupacion.get('Mujeres', 0).fillna(0)
                total_row = pd.DataFrame({
                    "Nivel": ["Total"],
                    "Hombres": [agrupacion["Hombres"].sum()],
                    "Mujeres": [agrupacion["Mujeres"].sum()],
                    "Total": [agrupacion["Total"].sum()]
                })
                agrupacion = pd.concat([agrupacion, total_row], ignore_index=True)
                agrupacion.to_excel(writer, sheet_name='Matricula Global', index=False)
                hojas_generadas.append('Matricula Global')

            for seleccion in indice:
                indice_dataframe = dataframe[dataframe['Indice'] == seleccion]
                if seleccion == 'Egresados':
                    if (len(indice_dataframe['Sexo'].unique()) >= 2):
                        agrupacion = indice_dataframe.groupby(['Nivel','Sexo'], as_index=False)['Datos'].sum()
                        agrupacion = agrupacion.pivot(index=['Nivel'], columns='Sexo', values='Datos').reset_index()
                        agrupacion['Total'] = agrupacion.get('Hombres', 0).fillna(0) + agrupacion.get('Mujeres', 0).fillna(0)

                        # Evita nombres de hoja inválidos
                        nombre_hoja = re.sub(r'[\[\]\*\?\/\\]', '', seleccion)[:31]
                        agrupacion.to_excel(writer, sheet_name=nombre_hoja, index=False)
                        hojas_generadas.append(nombre_hoja)
                else:
                    if len(indice_dataframe['Sexo'].unique()) >= 2 and len(indice_dataframe['Concepto'].unique()) >= 1:
                        agrupacion = indice_dataframe.groupby(['Nivel', 'Concepto', 'Sexo'], as_index=False)['Datos'].sum()
                        agrupacion = agrupacion.pivot(index=['Nivel', 'Concepto'], columns='Sexo', values='Datos').reset_index()
                        agrupacion['Total'] = agrupacion.get('Hombres', 0).fillna(0) + agrupacion.get('Mujeres', 0).fillna(0)

                        # Evita nombres de hoja inválidos
                        nombre_hoja = re.sub(r'[\[\]\*\?\/\\]', '', seleccion)[:31]
                        agrupacion.to_excel(writer, sheet_name=nombre_hoja, index=False)
                        hojas_generadas.append(nombre_hoja)
                

        # Aplicar estilos a los encabezados
        wb = load_workbook(ruta_salida)
        fill = PatternFill(start_color='5A1236', end_color='5A1236', fill_type='solid')
        font = Font(color='FFFFFF', bold=True, size=12)
        align = Alignment(horizontal='center', vertical='center')

        for hoja in hojas_generadas:
            ws = wb[hoja]
            for cell in ws[1]:
                cell.fill = fill
                cell.font = font
                cell.alignment = align

        orden = ['Medio Superior', 'Superior', 'Posgrado']
        ordenar_y_agrupar_columna_en_libro(wb, 'Nivel', orden)
        wb.save(ruta_salida)
        

    except Exception as e:
        print(f"Se produjo un error: {e}")