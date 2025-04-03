import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 📌 Crear DataFrame de ejemplo
df = pd.DataFrame({
    "Columna1": [10, 20, 30],
    "Columna2": [40, 50, 60]
})

# 📌 Guardar el DataFrame en un archivo Excel sin estilos
ruta_excel = "archivo_con_estilo.xlsx"
df.to_excel(ruta_excel, index=False)  # Guardar sin índice

# 📌 Cargar el archivo con openpyxl para agregar estilos
wb = load_workbook(ruta_excel)
ws = wb.active  # Obtener la hoja activa

# 📌 Definir estilo del encabezado
encabezado_style = {
    "font": Font(size=12, bold=True, color="FFFFFF"),
    "fill": PatternFill("solid", fgColor="5A1236"),
    "alignment": Alignment(horizontal="center", vertical="center")
}

# 📌 Aplicar estilo a los encabezados
for col_idx, celda in enumerate(ws[1], 1):  # Primera fila
    celda.font = encabezado_style["font"]
    celda.fill = encabezado_style["fill"]
    celda.alignment = encabezado_style["alignment"]

# 📌 Ajustar tamaño de las columnas automáticamente
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_length + 2

# 📌 Guardar cambios en el archivo
wb.save(ruta_excel)
